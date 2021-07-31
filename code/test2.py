import tkinter as tk
from tkinter import Message, filedialog
import tkinter.messagebox

import pandas as pd
import os
import openpyxl
import jieba
import re
import csv

from sklearn.feature_extraction.text import CountVectorizer
from sklearn.linear_model import RidgeClassifier
from sklearn.metrics import f1_score

from collections import Counter

filename = ''


class Application(tk.Frame):

    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.grid()
        self.createWidget()

    def upload_file(self):
        # 上传数据
        global filename
        filename= self.selectFile = tk.filedialog.askopenfilename() # askopenfilename 1次上传1个；askopenfilenames1次上传多个
        self.entry1.insert(0, self.selectFile)

    def createWidget(self):
        # 创建组件
        self.frm = tk.Frame(root)
        self.frm.grid(padx='20', pady='30')
        self.btn = tk.Button(self.frm, text='上传待分类数据', command=self.upload_file)
        self.btn.grid(row=0, column=0, ipadx='3', ipady='3', padx='10', pady='10')
        self.entry1 = tk.Entry(self.frm, width='30')
        self.entry1.grid(row=0, column=1)
        self.confirm = tk.Button(self.frm, text='确认', command= lambda: self.classificaton())
        self.confirm.grid(row=0, column=2)
        pass

    def classificaton(self):
        if filename == '':
            self.result = tk.messagebox.askokcancel(title='报错',message='请选择文件')
            self.entry1.delete(0, 'end')
        elif filename[-4:] != 'xlsx':
            self.result = tk.messagebox.askokcancel(title='报错', message='请选择xlsx文件')
            self.entry1.delete(0, 'end')
        else:
            # 调用停词表
            with open("data/hit_stopwords.txt", 'r', encoding='utf-8') as f:
                stopword_list = [w.strip('\n') for w in f.readlines()]

            # 测试数据
            filePath = filename

            wb = openpyxl.load_workbook(filePath)  # 打开excel文件

            ws = wb['类别']
            results = []

            # 清洗content 生成 list文件
            for i in range(2, ws.max_row + 1):
                print("正在处理第{}条数据, waiting".format(i))
                lists = []
                strs = ""
                strs = strs + ws.cell(i, 3).value + " " + ws.cell(i, 4).value

                # 去除中文标点符号
                strs = re.sub("[\s+\.\!\-\/_,$%^*()+\"\']+|[+——！，。？、~@#￥%……&*（）：”“]", "", strs)
                # 使用jieba进行分词，将文本分成词语列表
                words = jieba.lcut(strs)
                result = ""
                # 然后清除停用词语
                for word in words:
                    if word not in stopword_list:
                        result = result + word + " "

                lists.append(ws.cell(i, 1).value)
                lists.append(result)
                results.append(lists)

            # 清洗后的数据 写入csv

            header = ['Number', 'text']
            with open('data/test_set.csv', 'w', encoding='utf-8', newline='') as f:  # 解决空行的问题
                writer = csv.writer(f)
                writer.writerow(header)
                writer.writerows(results)
            f.close()

            # Count Vectors + RidgeClassifier

            train_df = pd.read_csv("data/train_set.csv", encoding='ANSI')
            test_df = pd.read_csv("data/test_set.csv", encoding="utf-8")

            vectorizer = CountVectorizer(max_features=3000)
            train_test = vectorizer.fit_transform(train_df['text'].values.astype('U'))
            test_test = vectorizer.fit_transform(test_df['text'].values.astype('U'))

            clf = RidgeClassifier()
            clf.fit(train_test, train_df['label'].values)

            # 预测集
            val_pred = clf.predict(train_test[10000:])

            # print(f1_score(train_df['label'].values[10000:], val_pred, average='macro'))
            # 测试集
            val_pred2 = clf.predict(test_test)

            filePath = "data/type.xlsx"
            wb = openpyxl.load_workbook(filePath)  # 打开excel文件

            ws = wb['类别']

            for i in range(2, ws.max_row + 1):
                ws.cell(row=i, column=2, value=str(val_pred2[i - 2]))

            wb.save("data/type.xlsx")



root = tk.Tk()
root.title('新闻文本分类')
root.geometry("400x100+200+300")
app = Application(master=root)

root.mainloop()
