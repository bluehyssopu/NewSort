import pandas as pd
import os
import openpyxl
import jieba
import re
import csv

from sklearn.feature_extraction.text import CountVectorizer
from sklearn.linear_model import RidgeClassifier
from sklearn.metrics import f1_score

from collections import Counter

# 调用停词表
with open("data/hit_stopwords.txt", 'r', encoding='utf-8') as f:
    stopword_list = [w.strip('\n') for w in f.readlines()]

# 测试数据
filePath = "data/type.xlsx"

wb = openpyxl.load_workbook(filePath)   # 打开excel文件

ws = wb['类别']
results = []

# 清洗content 生成 list文件
for i in range(2, ws.max_row+1):
    lists = []
    strs = ""
    strs = strs + ws.cell(i, 3).value + " " + ws.cell(i, 4).value

    # 去除中文标点符号
    strs = re.sub("[\s+\.\!\-\/_,$%^*()+\"\']+|[+——！，。？、~@#￥%……&*（）：”“]", "",strs)
    # 使用jieba进行分词，将文本分成词语列表
    words = jieba.lcut(strs)
    result = ""
    # 然后清除停用词语
    for word in words:
        if word not in stopword_list:
            result = result + word + " "


    lists.append(ws.cell(i, 1).value)
    lists.append(result)
    results.append(lists)

# 清洗后的数据 写入csv

header = ['Number', 'text']
with open('data/test_set.csv', 'w', encoding='utf-8', newline='') as f: # 解决空行的问题
    writer = csv.writer(f)
    writer.writerow(header)
    writer.writerows(results)
f.close()

# Count Vectors + RidgeClassifier

train_df = pd.read_csv("data/train_set.csv",encoding='ANSI')
test_df = pd.read_csv("data/test_set.csv", encoding="utf-8")

vectorizer = CountVectorizer(max_features=3000)
train_test = vectorizer.fit_transform(train_df['text'].values.astype('U'))
test_test = vectorizer.fit_transform(test_df['text'].values.astype('U'))

clf = RidgeClassifier()
clf.fit(train_test, train_df['label'].values)

# 预测集
val_pred = clf.predict(train_test[10000:])

# print(f1_score(train_df['label'].values[10000:], val_pred, average='macro'))
# 测试集
val_pred2 = clf.predict(test_test)


filePath = "data/type.xlsx"
wb = openpyxl.load_workbook(filePath)   # 打开excel文件

ws = wb['类别']

for i in range(2, ws.max_row+1):
    ws.cell(row=i, column=2, value = str(val_pred2[i-2]))

wb.save("data/type.xlsx")
















import pandas as pd
train_df = pd.read_csv('../rootData/train_copy.csv')

train_df['content']
for i in range(len(train_df)):
    strs = train_df['content'][i]
    #print(strs)
    if(i<5):
        print(len(str(strs)))   # object of type 'float' has no len() 强制转换 纺织报错
        print()

# 显示各新闻的长度分布 这里只是粗略显示 句号分割
train_df['text_len'] = train_df['content'].apply(lambda x: len(str(x).split('。')))
train_df['text_len'].describe()

import matplotlib.pyplot as plt
_ = plt.hist(train_df['text_len'], bins=240)
plt.xlabel('Text char count')
plt.title("Histogram of char count")


# 生成各类新闻数量的直方图
%matplotlib inline
plt.rcParams['font.sans-serif']=['SimHei']
plt.rcParams['axes.unicode_minus']=False

train_df['channelName'].value_counts().plot(kind='barh')
plt.title('News class count')
plt.xlabel("category")
# plt.savefig("label2.png")

# 写入所有content和title 至content.txt

train_df['content']
f = open("../rootData/content.txt","w",encoding='utf-8')
for i in range(len(train_df)):
    strs = train_df['content'][i]
    f.write(str(strs))
    #print(strs)
f.close()

# 统计content的词频
import jieba

file = open("../rootData/content.txt", "r", encoding='utf-8') #此处需打开txt格式且编码为UTF-8的文本
txt = file.read()
words = jieba.lcut(txt)      # 使用jieba进行分词，将文本分成词语列表

count = {}
for word in words:            #  使用 for 循环遍历每个词语并统计个数
    if len(word) < 2:          # 排除单个字的干扰，使得输出结果为词语
        continue
    else:
        count[word] = count.get(word, 0) + 1    
        # 如果字典里键为 word 的值存在，则返回键的值并加一，如果不存在键word，则返回0再加上1
        
exclude = ["可以", "一起", "这样"]  # 建立无关词语列表
for key in list(count.keys()):     # 遍历字典的所有键，即所有word
    if key in exclude:
        del count[key]                  #  删除字典中键为无关词语的键值对
        
list = list(count.items())         # 将字典的所有键值对转化为列表
list.sort(key=lambda x: x[1], reverse=True)     # 对列表按照词频从大到小的顺序排序

for i in range(20):  #   此处统计排名前五的单词，所以range(5)
    word, number = list[i]
    print("关键字：{:-<10}频次：{:+>8}".format(word, number))


# 导入扩展库 生成词云
import re           # 正则表达式库
import collections  # 词频统计库
import numpy as np  # numpy数据处理库
import jieba        # 结巴分词
import wordcloud    # 词云展示库
from PIL import Image # 图像处理库
import matplotlib.pyplot as plt # 图像展示库
from pylab import mpl           # 用于处理中文乱码

# 读取文件
fn = open('../rootData/content.txt', encoding='utf-8') # 打开文件并编码
string_data = fn.read() # 读出整个文件
fn.close() # 关闭文件

# 文本预处理
pattern = re.compile(u'\t|\n|\.|-|:|;|\)|\(|\?|\（|\）"') # 定义正则表达式匹配模式
string_data = re.sub(pattern, '', string_data) # 将符合模式的字符去除

# 文本分词
seg_list_exact = jieba.cut(string_data, cut_all = False) # 精确模式分词
object_list = []

# 加载 哈工大的停用词表 清洗用到
with open("../data/hit_stopwords.txt", 'r', encoding='utf-8') as f:
    remove_words = [w.strip('\n') for w in f.readlines()]

for word in seg_list_exact: # 循环读出每个分词
    if word not in remove_words: # 如果不在去除词库中
        object_list.append(word) # 分词追加到列表

# 词频统计
word_counts = collections.Counter(object_list) # 对分词做词频统计
word_counts_top10 = word_counts.most_common(10) # 获取前10最高频的词
print (word_counts_top10) # 输出检查

# 词频展示 
mask = np.array(Image.open('../rootData/TestImg.jpg')) # 定义词频背景
wc = wordcloud.WordCloud(
    font_path='C:/Windows/Fonts/SimHei.ttf', # 设置字体格式
    mask=mask, # 设置背景图
    max_words=200, # 最多显示词数
    max_font_size=100, # 字体最大值
    background_color='white'# 设置背景颜色,默认为黑色black
)

fig = plt.figure(figsize=(10, 8)) # 设置显示窗口大小
wc.generate_from_frequencies(word_counts) # 从字典生成词云
image_colors = wordcloud.ImageColorGenerator(mask) # 从背景图建立颜色方案
wc.recolor(color_func=image_colors) # 将词云颜色设置为背景图方案
mpl.rcParams['font.sans-serif'] = ['SimHei']
plt.suptitle('anysis') #这里设置中文可能会乱码，需要导包
plt.imshow(wc) # 显示词云
plt.axis('off') # 关闭坐标轴
plt.show() # 显示图像
wc.to_file('../rootData/result01.jpg') # 无白边保存图片
fig.savefig('../rootData/result02.jpg') # 有白边保存

# 查看词云 说明清洗有效

import jieba
import pandas as pd
import re

train_df = pd.read_csv('../rootData/train_copy.csv')
train_df['content']

dic = {'科技': 0,  '体育': 2, '娱乐': 3,  '军事': 4, '其他': 5, '教育': 6, '财经': 7, '汽车': 8, '游戏': 9, '房产': 10 }

# 读取停用词
with open("../data/hit_stopwords.txt", 'r', encoding='utf-8') as f:
    stopword_list = [w.strip('\n') for w in f.readlines()]

Endlist = []

for i in range(len(train_df['content'])):
    lists = []
    title = str(train_df['title'][i])
    strs = str(train_df['content'][i]) + title
    labels = str(train_df['channelName'][i])
    if(labels in dic) :
        labels = dic[labels]
    
    # 去除中文标点符号
    strs = re.sub("[\s+\.\!\-\/_,$%^*()+\"\']+|[+——！，。？、~@#￥%……&*（）：”“]", "",strs)
    # 使用jieba进行分词，将文本分成词语列表
    words = jieba.lcut(strs)
    result = ""
    # 然后清除停用词语
    for word in words:
        if word not in stopword_list:
            result = result + " " + word
    # print(result)
    # print(type(result))
    # print()

    lists.append(labels)
    lists.append(result)
    Endlist.append(lists)
    # 最终是一个二维矩阵 方便 pandas 操作
print(Endlist[:10])


import csv

header = ['label', 'text']
with open('../rootData/train_set.csv', 'w', encoding="utf-8", newline='') as f: # 解决空行的问题
    writer = csv.writer(f)
    writer.writerow(header)
    writer.writerows(Endlist)
f.close()


# 处理测试集 xlsx中的脏数据

import openpyxl
import jieba
import re

# 加载 哈工大的停用词表 清洗用到
with open("../data/hit_stopwords.txt", 'r', encoding='utf-8') as f:
    stopword_list = [w.strip('\n') for w in f.readlines()]

filePath = "../data/type.xlsx"
wb = openpyxl.load_workbook(filePath)   # 打开excel文件
ws = wb['类别']     # 定位 sheet
results = []

# 这里把title 和 conten结合成一个str
for i in range(2, ws.max_row+1):
    lists = []
    strs = ""
    strs = strs + ws.cell(i, 3).value + " " + ws.cell(i, 4).value

    # 去除中文标点符号
    strs = re.sub("[\s+\.\!\-\/_,$%^*()+\"\']+|[+——！，。？、~@#￥%……&*（）：”“]", "",strs)
    # 使用jieba进行分词，将文本分成词语列表
    words = jieba.lcut(strs)
    result = ""
    # 然后清除停用词语
    for word in words:
        if word not in stopword_list:
            result = result + word + " "


    lists.append(ws.cell(i, 1).value)
    lists.append(result)
    results.append(lists)

# 得到一个list 然后写入csv
# print(results)


# 生成 csv文件
import csv

header = ['Number', 'text']
with open('../data/test_set.csv', 'w', encoding='utf-8', newline='') as f: # 解决空行的问题
    writer = csv.writer(f)
    writer.writerow(header)
    writer.writerows(results)

f.close()       # 记得关闭文件


# Count Vectors + RidgeClassifier

import pandas as pd

from sklearn.feature_extraction.text import CountVectorizer
from sklearn.linear_model import RidgeClassifier
from sklearn.metrics import f1_score

train_df = pd.read_csv("../data/train_set.csv",encoding='ANSI')
test_df = pd.read_csv("../data/test_set.csv", encoding="utf-8")

vectorizer = CountVectorizer(max_features=3000)
# 定位 训练列
train_test = vectorizer.fit_transform(train_df['text'].values.astype('U'))
test_test = vectorizer.fit_transform(test_df['text'].values.astype('U'))

clf = RidgeClassifier()
# 这里可以设置 训练集的范文围 如 train_test[:10000] 训练集的 70%训练 30%预测
clf.fit(train_test, train_df['label'].values)

# 得到模型的得分
val_pred = clf.predict(train_test[10000:])

# 输出得分
print(f1_score(train_df['label'].values[10000:], val_pred, average='macro'))

# 测试集 预测
val_pred2 = clf.predict(test_test)

# 写入预测值至xlsx
filePath = "../data/type.xlsx"
wb = openpyxl.load_workbook(filePath)   # 打开excel文件

ws = wb['类别']

for i in range(2, ws.max_row+1):
    ws.cell(row=i, column=2, value = str(val_pred2[i-2]))

wb.save("../data/type.xlsx")


# 打包
# pyinstaller -p D:\\code\\Python_work\\NewSortEnd\\code -F test2.py --hidden-import sklearn.utils._weight_vector

# 读取文件
import pandas as pd

filename = "../data/train_set.csv"
train_df = pd.read_csv(filename, encoding="ANSI")


# 显示前五行
train_df.head()


# 简单显示文本长度
train_df['text_len'] = train_df['text'].apply(lambda x: len(str(x).split(' ')))
train_df['text_len'].describe()



# 绘制分布图
import matplotlib.pyplot as plt
_ = plt.hist(train_df['text_len'], bins=200)
plt.xlabel('Text char count')
plt.title("Histogram of char count")


%matplotlib inline
plt.rcParams['font.sans-serif']=['SimHei']
plt.rcParams['axes.unicode_minus']=False

train_df['label'].value_counts().plot(kind='barh')
plt.title('News class count')
plt.xlabel("category")


# 统计词频
from collections import Counter
all_lines = ' '.join(str(train_df['text']))
word_count = Counter(all_lines.split(" "))
word_count = sorted(word_count.items(), key=lambda d:d[1], reverse = True)

print(len(word_count))

print(word_count[0:10])


# 词袋

from collections import Counter
train_df['text_unique'] = train_df['text'].apply(lambda x: ' '.join(list(set(str(x).split(' ')))))
all_lines = ' '.join(list(train_df['text_unique']))
word_count = Counter(all_lines.split(" "))
word_count = sorted(word_count.items(), key=lambda d:int(d[1]), reverse = True)

print(word_count[0])
print(word_count[1])
print(word_count[2])
print(word_count[3])
print(word_count[4])
print(word_count[5])


# Count Vectors + RidgeClassifier
# 词袋 + 特征值提取
import pandas as pd

from sklearn.feature_extraction.text import CountVectorizer
from sklearn.linear_model import RidgeClassifier
from sklearn.metrics import f1_score

train_df = pd.read_csv('../data/train_set.csv',encoding='ANSI')

vectorizer = CountVectorizer(max_features=3000)
train_test = vectorizer.fit_transform(train_df['text'].values.astype('U'))

clf = RidgeClassifier()
clf.fit(train_test[:10000], train_df['label'].values[:10000])

val_pred = clf.predict(train_test[10000:])

print(f1_score(train_df['label'].values[10000:], val_pred, average='macro'))
print(type(val_pred))

# 0.10343472451465849
# 0.10357460616527975 
# 一定要打乱训练的顺序
# 0.8491012463815872
# 0.8207100395680607



# 朴素贝叶斯
# TF-IDF +  RidgeClassifier

import pandas as pd

from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import RidgeClassifier
from sklearn.metrics import f1_score

train_df = pd.read_csv('../data/train_set.csv', encoding="ANSI")

tfidf = TfidfVectorizer(ngram_range=(1,3), max_features=3000)
train_test = tfidf.fit_transform(train_df['text'])

clf = RidgeClassifier()
clf.fit(train_test[:10000], train_df['label'].values[:10000])

val_pred = clf.predict(train_test[10000:])
print(f1_score(train_df['label'].values[10000:], val_pred, average='macro'))
# 0.10658403152377673
# 0.10661056926117167
# 0.8404218282357144
# 0.8602994126998214


import tkinter as tk
from tkinter import Message, filedialog
import tkinter.messagebox

import pandas as pd
import os
import openpyxl
import jieba
import re
import csv

from sklearn.feature_extraction.text import CountVectorizer
from sklearn.linear_model import RidgeClassifier
from sklearn.metrics import f1_score

from collections import Counter

filename = ''


class Application(tk.Frame):

    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.grid()
        self.createWidget()

    def upload_file(self):
        # 上传数据
        global filename
        filename= self.selectFile = tk.filedialog.askopenfilename() # askopenfilename 1次上传1个；askopenfilenames1次上传多个
        self.entry1.insert(0, self.selectFile)

    def createWidget(self):
        # 创建组件
        self.frm = tk.Frame(root)
        self.frm.grid(padx='20', pady='30')
        self.btn = tk.Button(self.frm, text='上传待分类数据', command=self.upload_file)
        self.btn.grid(row=0, column=0, ipadx='3', ipady='3', padx='10', pady='10')
        self.entry1 = tk.Entry(self.frm, width='30')
        self.entry1.grid(row=0, column=1)
        self.confirm = tk.Button(self.frm, text='确认', command= lambda: self.classificaton())
        self.confirm.grid(row=0, column=2)
        pass

    def classificaton(self):
        if filename == '':
            self.result = tk.messagebox.askokcancel(title='报错',message='请选择文件')
            self.entry1.delete(0, 'end')
        elif filename[-4:] != 'xlsx':
            self.result = tk.messagebox.askokcancel(title='报错', message='请选择xlsx文件')
            self.entry1.delete(0, 'end')
        else:
            # 调用停词表
            with open("data/hit_stopwords.txt", 'r', encoding='utf-8') as f:
                stopword_list = [w.strip('\n') for w in f.readlines()]

            # 测试数据
            filePath = filename

            wb = openpyxl.load_workbook(filePath)  # 打开excel文件

            ws = wb['类别']
            results = []

            # 清洗content 生成 list文件
            for i in range(2, ws.max_row + 1):
                print("正在处理第{}条数据, waiting".format(i))
                lists = []
                strs = ""
                strs = strs + ws.cell(i, 3).value + " " + ws.cell(i, 4).value

                # 去除中文标点符号
                strs = re.sub("[\s+\.\!\-\/_,$%^*()+\"\']+|[+——！，。？、~@#￥%……&*（）：”“]", "", strs)
                # 使用jieba进行分词，将文本分成词语列表
                words = jieba.lcut(strs)
                result = ""
                # 然后清除停用词语
                for word in words:
                    if word not in stopword_list:
                        result = result + word + " "

                lists.append(ws.cell(i, 1).value)
                lists.append(result)
                results.append(lists)

            # 清洗后的数据 写入csv

            header = ['Number', 'text']
            with open('data/test_set.csv', 'w', encoding='utf-8', newline='') as f:  # 解决空行的问题
                writer = csv.writer(f)
                writer.writerow(header)
                writer.writerows(results)
            f.close()

            # Count Vectors + RidgeClassifier

            train_df = pd.read_csv("data/train_set.csv", encoding='ANSI')
            test_df = pd.read_csv("data/test_set.csv", encoding="utf-8")

            vectorizer = CountVectorizer(max_features=3000)
            train_test = vectorizer.fit_transform(train_df['text'].values.astype('U'))
            test_test = vectorizer.fit_transform(test_df['text'].values.astype('U'))

            clf = RidgeClassifier()
            clf.fit(train_test, train_df['label'].values)

            # 预测集
            val_pred = clf.predict(train_test[10000:])

            # print(f1_score(train_df['label'].values[10000:], val_pred, average='macro'))
            # 测试集
            val_pred2 = clf.predict(test_test)

            filePath = "data/type.xlsx"
            wb = openpyxl.load_workbook(filePath)  # 打开excel文件

            ws = wb['类别']

            for i in range(2, ws.max_row + 1):
                ws.cell(row=i, column=2, value=str(val_pred2[i - 2]))

            wb.save("data/type.xlsx")



root = tk.Tk()
root.title('新闻文本分类')
root.geometry("400x100+200+300")
app = Application(master=root)

root.mainloop()
